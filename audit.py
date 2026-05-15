"""Backlink audit core logic.

Reads an Ahrefs-style backlinks XLSX (subdomains export) and produces a
styled two-sheet report similar to chi_tiet_anchor_backlink_reno15.xlsx.
"""
from __future__ import annotations

import json
import re
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from typing import Iterable
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


AHREFS_BACKLINKS_ENDPOINT = "https://api.ahrefs.com/v3/site-explorer/all-backlinks"
AHREFS_SELECT_FIELDS = (
    "url_from,url_to,anchor,title,domain_rating_source,"
    "first_seen_link,is_dofollow,is_spam"
)


class AhrefsAPIError(Exception):
    """Raised when the Ahrefs API returns an error response."""

    def __init__(self, message: str, status: int | None = None):
        super().__init__(message)
        self.status = status


# --- Styling constants (matches the reno15 reference workbook) ---------------

TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
LEGEND_FILL = PatternFill("solid", fgColor="FFF2CC")

DR_HIGH_FILL = PatternFill("solid", fgColor="E2EFDA")   # DR >= 50: green
DR_MID_FILL = PatternFill("solid", fgColor="FFF2CC")    # DR 20-49: yellow
DR_ZEBRA_FILL = PatternFill("solid", fgColor="F2F7FB")  # DR < 20 alt rows

WHITE_FONT = Font(name="Helvetica", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Helvetica", color="FFFFFF", bold=True, size=14)
BODY_FONT = Font(name="Helvetica", size=11)
BODY_BOLD = Font(name="Helvetica", size=11, bold=True)

THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Anchor / referring page hints that indicate Black-Hat / spam content.
SPAM_HINTS = re.compile(
    r"(black\s*hat|telegram\s*[:@]|↑↑↑|casino|bet\b|judi|porn|viagra|"
    r"cialis|gambling|crypto\s*pump|fast\s*ranking|backlink\s*service)",
    re.IGNORECASE,
)


# --- Helpers -----------------------------------------------------------------


def _domain_of(url: str) -> str:
    """Return a clean registrable host for a URL (strip leading www.)."""
    if not url:
        return ""
    try:
        netloc = urlparse(str(url)).netloc.lower()
    except ValueError:
        return ""
    return netloc[4:] if netloc.startswith("www.") else netloc


def _path_of(url: str) -> str:
    """Return path component minus the leading slash (e.g. 'dtdd/oppo-reno15-...')."""
    if not url:
        return ""
    try:
        parsed = urlparse(str(url))
    except ValueError:
        return str(url)
    path = (parsed.path or "").lstrip("/")
    return path or str(url)


def _is_spammy(row: dict) -> bool:
    """True when a backlink row looks like Black-Hat SEO."""
    if str(row.get("Is spam", "")).strip().lower() in {"true", "1", "yes"}:
        return True
    anchor = str(row.get("Anchor", "") or "")
    ref_title = str(row.get("Referring page title", "") or "")
    if SPAM_HINTS.search(anchor) or SPAM_HINTS.search(ref_title):
        return True
    return False


def _fmt_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value)
    # Ahrefs export often uses "2026-05-08 15:51:59" — keep date only.
    return s.split(" ", 1)[0]


def _to_number(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


# --- Reading -----------------------------------------------------------------


def read_backlinks(path: str) -> list[dict]:
    """Load an Ahrefs backlinks XLSX (first sheet) into a list of dict rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []
    records: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        record = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        if not any(record.values()):
            continue
        records.append(record)
    wb.close()
    return records


def fetch_from_ahrefs(domain: str, api_key: str,
                      mode: str = "subdomains",
                      limit: int = 1000,
                      history: str = "all_time",
                      timeout: int = 60) -> list[dict]:
    """Fetch backlinks via Ahrefs API v3 and map fields to XLSX-style keys.

    Returns rows with the same keys read_backlinks() produces, so the rest of
    the audit pipeline can consume them unchanged.
    """
    target = (domain or "").strip()
    if not target:
        raise ValueError("Domain trống.")
    if not (api_key or "").strip():
        raise ValueError("API key trống.")

    # Strip scheme / trailing slash so people can paste either form.
    if "://" in target:
        target = urlparse(target).netloc or target
    target = target.strip("/")

    params = {
        "target": target,
        "mode": mode,
        "protocol": "both",
        "select": AHREFS_SELECT_FIELDS,
        "limit": str(int(limit)),
        "history": history,
        "output": "json",
    }
    url = f"{AHREFS_BACKLINKS_ENDPOINT}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {api_key.strip()}",
        "Accept": "application/json",
    })

    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            detail = exc.read().decode("utf-8", errors="replace")[:400]
        except Exception:
            pass
        if exc.code == 401:
            raise AhrefsAPIError(
                "API key không hợp lệ hoặc đã hết hạn (HTTP 401).", 401)
        if exc.code == 403:
            raise AhrefsAPIError(
                "Subscription của bạn không có quyền dùng endpoint này "
                "(HTTP 403). Cần plan Standard trở lên.", 403)
        if exc.code == 429:
            raise AhrefsAPIError(
                "Đã vượt rate limit của Ahrefs API (HTTP 429). Thử lại sau.",
                429)
        raise AhrefsAPIError(
            f"Ahrefs API trả lỗi HTTP {exc.code}: {detail}", exc.code)
    except urllib.error.URLError as exc:
        raise AhrefsAPIError(f"Không kết nối được Ahrefs API: {exc.reason}")

    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        raise AhrefsAPIError("Ahrefs trả về response không phải JSON.")

    raw_rows = data.get("backlinks") or []
    rows: list[dict] = []
    for r in raw_rows:
        rows.append({
            "Target URL": r.get("url_to") or "",
            "Anchor": r.get("anchor") or "",
            "Referring page URL": r.get("url_from") or "",
            "Referring page title": r.get("title") or "",
            "Domain rating": r.get("domain_rating_source"),
            "First seen": r.get("first_seen_link") or "",
            "Is spam": str(r.get("is_spam", "")).lower(),
        })
    return rows


# --- Audit pipeline ----------------------------------------------------------


def filter_rows(rows: Iterable[dict], keyword: str) -> list[dict]:
    """Keep rows whose Target URL contains the keyword and that aren't spam."""
    kw = (keyword or "").strip().lower()
    out: list[dict] = []
    for r in rows:
        target = str(r.get("Target URL", "") or "")
        if kw and kw not in target.lower():
            continue
        if not target:
            continue
        if _is_spammy(r):
            continue
        out.append(r)
    return out


def build_detail(rows: list[dict]) -> dict[str, list[dict]]:
    """Group filtered rows by full target URL, sorted by DR desc then date desc."""
    groups: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        target = str(r.get("Target URL", "") or "").split("?", 1)[0]
        anchor = str(r.get("Anchor", "") or "").strip() or "(không có anchor)"
        groups[target].append({
            "target_full": target,
            "target_path": _path_of(target),
            "anchor": anchor,
            "ref_url": str(r.get("Referring page URL", "") or ""),
            "dr": _to_number(r.get("Domain rating")) or 0,
            "first_seen": _fmt_date(r.get("First seen")),
        })
    for target, items in groups.items():
        items.sort(key=lambda x: (-(x["dr"] or 0), x["first_seen"]), reverse=False)
        # We want DR desc, date desc: sort by date desc inside the negative-DR sort.
        items.sort(key=lambda x: (-(x["dr"] or 0), -_date_sort_key(x["first_seen"])))
    return dict(sorted(groups.items()))


def _date_sort_key(s: str) -> int:
    if not s:
        return 0
    try:
        return int(datetime.strptime(s, "%Y-%m-%d").timestamp())
    except ValueError:
        return 0


def build_domain_summary(detail: dict[str, list[dict]]) -> dict[str, list[dict]]:
    """For each target URL, aggregate by source domain with link count + max DR."""
    summary: dict[str, list[dict]] = {}
    for target, items in detail.items():
        per_domain: dict[str, dict] = {}
        for it in items:
            dom = _domain_of(it["ref_url"])
            if not dom:
                continue
            row = per_domain.setdefault(dom, {"domain": dom, "count": 0, "dr_max": 0})
            row["count"] += 1
            if it["dr"] and it["dr"] > row["dr_max"]:
                row["dr_max"] = it["dr"]
        rows = list(per_domain.values())
        rows.sort(key=lambda x: (-x["count"], -x["dr_max"]))
        summary[target] = rows
    return summary


# --- Writing -----------------------------------------------------------------


def _apply_dr_fill(ws, row_idx: int, cols: int, dr: float, zebra: bool, mode: str):
    """Paint a data row's background based on Domain Rating.

    mode: 'detail' — green when DR>=50, otherwise zebra striping.
          'summary' — green when DR>=50, yellow 20<=DR<50, else zebra.
    """
    if dr is None:
        dr = 0
    if dr >= 50:
        fill = DR_HIGH_FILL
    elif mode == "summary" and dr >= 20:
        fill = DR_MID_FILL
    else:
        fill = DR_ZEBRA_FILL if zebra else None
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        if fill is not None:
            cell.fill = fill
        cell.border = CELL_BORDER
        cell.font = BODY_BOLD if c == (5 if mode == "detail" else 5) else BODY_FONT
        cell.alignment = CENTER if c in {1, 5, 6} else LEFT


def _write_detail_sheet(ws, detail: dict[str, list[dict]], product_label: str):
    cols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.cell(row=1, column=1).value = (
        f"CHI TIẾT ANCHOR TEXT VÀ NGUỒN BACKLINK THEO TỪNG URL {product_label.upper()}"
    )
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws.cell(row=2, column=1).value = (
        "Chỉ bao gồm backlink chất lượng (đã lọc bỏ Black Hat SEO)  |  "
        "Màu xanh = DR ≥ 50"
    )
    ws.cell(row=2, column=1).fill = LEGEND_FILL
    ws.cell(row=2, column=1).font = BODY_BOLD
    ws.cell(row=2, column=1).alignment = CENTER

    headers = ["STT", "Target URL (TGDĐ)", "Anchor Text",
               "URL nguồn backlink", "Domain Rating", "Ngày phát hiện"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = CELL_BORDER

    widths = [5, 55, 35, 60, 15, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    for target, items in detail.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1)
        cell.value = f"📌  {target}   ({len(items)} backlink chất lượng)"
        cell.fill = GROUP_FILL
        cell.font = WHITE_FONT
        cell.alignment = LEFT
        cell.border = CELL_BORDER
        row += 1
        for i, it in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=it["target_path"])
            ws.cell(row=row, column=3, value=it["anchor"])
            ws.cell(row=row, column=4, value=it["ref_url"])
            ws.cell(row=row, column=5, value=int(it["dr"]) if it["dr"] and it["dr"] == int(it["dr"]) else it["dr"])
            ws.cell(row=row, column=6, value=it["first_seen"])
            _apply_dr_fill(ws, row, cols, it["dr"], zebra=(i % 2 == 0), mode="detail")
            row += 1

    ws.freeze_panes = "A4"


def _write_summary_sheet(ws, summary: dict[str, list[dict]], product_label: str):
    cols = 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.cell(row=1, column=1).value = (
        f"THỐNG KÊ DOMAIN ĐANG ĐI BACKLINK VÀO TỪNG URL {product_label.upper()}"
    )
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws.cell(row=2, column=1).value = (
        "Chỉ bao gồm backlink chất lượng (đã lọc Black Hat)  |  "
        "Xanh lá = DR ≥ 50  |  Vàng = DR 20–49"
    )
    ws.cell(row=2, column=1).fill = LEGEND_FILL
    ws.cell(row=2, column=1).font = BODY_BOLD
    ws.cell(row=2, column=1).alignment = CENTER

    headers = ["STT", "Target URL (TGDĐ)", "Domain nguồn", "Số link", "DR (max)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = CELL_BORDER

    widths = [5, 60, 28, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    for target, items in summary.items():
        target_path = _path_of(target)
        total_links = sum(it["count"] for it in items)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1)
        cell.value = f"📌  {target_path}   →  {total_links} link  từ  {len(items)} domain"
        cell.fill = GROUP_FILL
        cell.font = WHITE_FONT
        cell.alignment = LEFT
        cell.border = CELL_BORDER
        row += 1
        for i, it in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=target_path)
            ws.cell(row=row, column=3, value=it["domain"])
            ws.cell(row=row, column=4, value=it["count"])
            ws.cell(row=row, column=5, value=int(it["dr_max"]) if it["dr_max"] == int(it["dr_max"]) else it["dr_max"])
            _apply_dr_fill(ws, row, cols, it["dr_max"], zebra=(i % 2 == 0), mode="summary")
            row += 1

    ws.freeze_panes = "A4"


def _build_report(rows: list[dict], keyword: str, output_path: str,
                  product_label: str | None) -> dict:
    """Filter, group, and write the styled workbook. Shared by both audit modes.

    Returns a dict with both `stats` and `preview` so the UI can render a
    table inline (preview) and still offer the .xlsx download.
    """
    kept = filter_rows(rows, keyword)
    detail = build_detail(kept)
    summary = build_domain_summary(detail)

    label = (product_label or keyword or "report").strip() or "report"

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Chi Tiết URL vs Anchor"
    _write_detail_sheet(ws1, detail, label)

    ws2 = wb.create_sheet("📊 Domain vs URL")
    _write_summary_sheet(ws2, summary, label)

    wb.save(output_path)

    total_links = sum(len(v) for v in detail.values())
    total_domains = sum(len(v) for v in summary.values())

    detail_preview = [
        {
            "target": target,
            "target_path": _path_of(target),
            "count": len(items),
            "rows": [
                {
                    "stt": i + 1,
                    "anchor": it["anchor"],
                    "ref_url": it["ref_url"],
                    "dr": it["dr"],
                    "first_seen": it["first_seen"],
                }
                for i, it in enumerate(items)
            ],
        }
        for target, items in detail.items()
    ]
    summary_preview = [
        {
            "target": target,
            "target_path": _path_of(target),
            "total_links": sum(it["count"] for it in items),
            "domain_count": len(items),
            "rows": [
                {
                    "stt": i + 1,
                    "domain": it["domain"],
                    "count": it["count"],
                    "dr_max": it["dr_max"],
                }
                for i, it in enumerate(items)
            ],
        }
        for target, items in summary.items()
    ]

    return {
        "input_rows": len(rows),
        "kept_rows": len(kept),
        "target_urls": len(detail),
        "total_links": total_links,
        "total_domain_rows": total_domains,
        "preview": {
            "detail": detail_preview,
            "summary": summary_preview,
        },
    }


def run_audit(backlinks_path: str, keyword: str, output_path: str,
              product_label: str | None = None) -> dict:
    """Run the full audit from a local XLSX file."""
    rows = read_backlinks(backlinks_path)
    if not rows:
        raise ValueError("Không đọc được dữ liệu từ file backlinks.")
    return _build_report(rows, keyword, output_path, product_label)


def run_audit_from_ahrefs(domain: str, api_key: str, keyword: str,
                          output_path: str,
                          product_label: str | None = None,
                          limit: int = 1000,
                          mode: str = "subdomains") -> dict:
    """Fetch backlinks via Ahrefs API, then build the styled report."""
    rows = fetch_from_ahrefs(domain, api_key, mode=mode, limit=limit)
    if not rows:
        raise ValueError(
            "Ahrefs không trả về backlink nào cho domain này "
            "(kiểm tra lại domain hoặc subscription)."
        )
    return _build_report(rows, keyword, output_path, product_label)
